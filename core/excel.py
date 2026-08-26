"""Shared Excel reading and text conversion helpers."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
import re
from zipfile import BadZipFile

from .text_tools import decode_separator


SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
FORMULA_MODES = ("缓存值", "公式文本")
ORIENTATIONS = ("按行", "按列")
READ_DIRECTIONS = ("单行", "单列")
_CELL_RANGE_RE = re.compile(
    r"^(?P<start>[A-Za-z]+[1-9][0-9]*)(?::(?P<end>[A-Za-z]+[1-9][0-9]*))?$"
)
_CACHED_FORMULA_WITHOUT_VALUE = object()


@dataclass(frozen=True)
class ExcelTable:
    """Normalized rectangular worksheet values plus workbook metadata."""

    rows: tuple[tuple[str, ...], ...]
    sheet_names: tuple[str, ...]
    selected_sheet: str

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.rows[0]) if self.rows else 0


def _format_value(value) -> str:
    if value is None or value is _CACHED_FORMULA_WITHOUT_VALUE:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        if value.time() == time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def _is_empty(value) -> bool:
    return value is None or value == ""


def _restore_uncached_formula_cells(
    rows: list[list], formula_rows: list[list]
) -> list[list]:
    """Keep formula cells whose cached result was not saved from being cropped."""

    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            formula_value = formula_rows[row_index][column_index]
            if value is None and isinstance(formula_value, str) and formula_value.startswith("="):
                row[column_index] = _CACHED_FORMULA_WITHOUT_VALUE
    return rows


def _crop_outer_empty(rows: list[list]) -> list[list]:
    """Remove only fully empty rows/columns around a rectangular table."""

    while rows and all(_is_empty(value) for value in rows[0]):
        rows.pop(0)
    while rows and all(_is_empty(value) for value in rows[-1]):
        rows.pop()
    if not rows:
        return []

    left = 0
    width = len(rows[0])
    while left < width and all(_is_empty(row[left]) for row in rows):
        left += 1
    right = width
    while right > left and all(_is_empty(row[right - 1]) for row in rows):
        right -= 1
    return [row[left:right] for row in rows]


def _resolve_excel_path(excel_path: str) -> Path:
    raw_path = str(excel_path).strip().strip('"')
    if not raw_path:
        raise ValueError("Excel 路径不能为空。")
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        try:
            import folder_paths

            path = Path(folder_paths.get_annotated_filepath(raw_path))
        except (ImportError, AttributeError, TypeError, ValueError):
            raise ValueError("Excel 路径必须是绝对路径，或是通过节点选择的输入文件。") from None
    if path.suffix.lower() not in SUPPORTED_EXCEL_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_EXCEL_EXTENSIONS))
        raise ValueError(f"不支持的 Excel 格式：{path.suffix or '无扩展名'}；仅支持 {supported}。")
    if not path.exists():
        raise ValueError(f"找不到 Excel 文件：{path}")
    if not path.is_file():
        raise ValueError(f"Excel 路径不是文件：{path}")
    return path


def inspect_excel_workbook(excel_path: str) -> dict:
    """Return lightweight workbook metadata for the interactive node controls."""

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise RuntimeError(
            "缺少 Excel 读取依赖 openpyxl。请在 ComfyUI 的 Python 环境中安装 requirements.txt。"
        ) from exc

    path = _resolve_excel_path(excel_path)
    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, OSError, KeyError, ValueError, BadZipFile, EOFError) as exc:
        raise ValueError(f"无法读取 Excel 文件，文件可能已损坏或受密码保护：{path}") from exc

    try:
        from openpyxl.utils.cell import get_column_letter

        sheets = []
        for worksheet in workbook.worksheets:
            min_row = None
            min_column = None
            max_row = 0
            max_column = 0
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    if _is_empty(value):
                        continue
                    min_row = row_index if min_row is None else min(min_row, row_index)
                    min_column = column_index if min_column is None else min(min_column, column_index)
                    max_row = max(max_row, row_index)
                    max_column = max(max_column, column_index)
            if min_row is None or min_column is None:
                effective_range = "空工作表"
                effective_rows = 0
                effective_columns = 0
            else:
                effective_range = (
                    f"{get_column_letter(min_column)}{min_row}:"
                    f"{get_column_letter(max_column)}{max_row}"
                )
                effective_rows = max_row - min_row + 1
                effective_columns = max_column - min_column + 1
            sheets.append(
                {
                    "name": worksheet.title,
                    "range": effective_range,
                    "row_count": effective_rows,
                    "column_count": effective_columns,
                }
            )
        if not sheets:
            raise ValueError("Excel 文件中没有可读取的工作表。")
        return {"file_name": path.name, "sheets": sheets}
    finally:
        workbook.close()


def list_excel_input_files() -> list[dict[str, str]]:
    """List uploaded Excel workbooks from ComfyUI's dedicated input folder."""

    try:
        import folder_paths

        input_root = Path(folder_paths.get_input_directory()).resolve()
    except (ImportError, AttributeError, OSError):
        return []
    excel_root = (input_root / "TUT_Nodes" / "excel").resolve()
    if not excel_root.is_dir() or input_root not in excel_root.parents:
        return []
    try:
        paths = sorted(
            (
                path
                for path in excel_root.rglob("*")
                if path.is_file() and path.suffix.lower() in SUPPORTED_EXCEL_EXTENSIONS
            ),
            key=lambda path: path.relative_to(input_root).as_posix().lower(),
        )
    except OSError:
        return []
    return [
        {
            "name": path.name,
            "token": path.relative_to(input_root).as_posix(),
        }
        for path in paths
    ]


def pick_local_excel_file() -> dict[str, str] | None:
    """Open the server machine's native file picker without copying the workbook."""

    root = None
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        try:
            root.attributes("-topmost", True)
        except tk.TclError:
            pass
        root.update_idletasks()
        selected = filedialog.askopenfilename(
            parent=root,
            title="选择 Excel 文件",
            filetypes=(
                ("Excel 工作簿", "*.xlsx *.xlsm"),
                ("所有文件", "*.*"),
            ),
        )
    except Exception as exc:
        raise RuntimeError(f"无法打开本机 Excel 文件选择器：{exc}") from exc
    finally:
        if root is not None:
            try:
                root.destroy()
            except Exception:
                pass
    if not selected:
        return None
    path = _resolve_excel_path(selected)
    return {"path": str(path.resolve()), "name": path.name}


def excel_file_fingerprint(excel_path: str) -> str:
    """Return a cache key that changes when the selected file is replaced."""

    try:
        path = _resolve_excel_path(excel_path)
        stat = path.stat()
        return f"{path.resolve()}|{stat.st_mtime_ns}|{stat.st_size}"
    except (OSError, ValueError) as exc:
        return f"invalid|{excel_path}|{type(exc).__name__}|{exc}"


def _range_bounds(cell_range: str):
    value = str(cell_range).strip().replace("$", "")
    if not value:
        return None
    match = _CELL_RANGE_RE.fullmatch(value)
    if not match:
        raise ValueError("单元格范围格式无效，请使用 A1:F100 或 A1。")
    normalized = match.group("start")
    if match.group("end"):
        normalized += f":{match.group('end')}"
    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(normalized)
    except Exception as exc:
        raise ValueError(f"单元格范围无效：{cell_range}") from exc
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"单元格范围起点不能晚于终点：{cell_range}")
    return min_col, min_row, max_col, max_row


def _non_negative_int(value, label: str) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须是大于等于 0 的整数：{value!r}") from exc
    if parsed < 0:
        raise ValueError(f"{label}不能小于 0。")
    return parsed


def slice_excel_table(
    table: ExcelTable,
    skip_rows: int = 0,
    skip_columns: int = 0,
    orientation: str | None = None,
    max_items: int = 0,
) -> ExcelTable:
    """Skip table edges and optionally limit the resulting row/column groups."""

    if not isinstance(table, ExcelTable):
        raise ValueError("Excel 数据无效，请连接“TUT_读取Excel”节点。")
    skip_rows = _non_negative_int(skip_rows, "跳过行数")
    skip_columns = _non_negative_int(skip_columns, "跳过列数")
    max_items = _non_negative_int(max_items, "读取数量上限")
    if orientation is not None and orientation not in ORIENTATIONS:
        raise ValueError(f"未知组织方式：{orientation}")

    rows = table.rows[skip_rows:]
    rows = tuple(tuple(row[skip_columns:]) for row in rows)
    if not rows or not rows[0]:
        raise ValueError("跳过前 N 行或列后没有剩余数据。")

    if max_items > 0:
        if orientation == "按列":
            rows = tuple(tuple(row[:max_items]) for row in rows)
        else:
            rows = rows[:max_items]
    if not rows or not rows[0]:
        raise ValueError("应用读取数量上限后没有剩余数据。")
    return ExcelTable(rows, table.sheet_names, table.selected_sheet)


def select_excel_line(
    table: ExcelTable,
    read_direction: str,
    index: int,
    skip_items: int = 0,
    max_items: int = 0,
) -> tuple[str, ...]:
    """Select one zero-based row or column, then skip and limit its cells."""

    if not isinstance(table, ExcelTable):
        raise ValueError("Excel 数据无效，请连接“TUT_读取Excel”节点。")
    if read_direction not in READ_DIRECTIONS:
        raise ValueError(f"未知读取方向：{read_direction}")
    index = _non_negative_int(index, "编号")
    skip_items = _non_negative_int(skip_items, "跳过项目数")
    max_items = _non_negative_int(max_items, "读取数量上限")

    available = table.row_count if read_direction == "单行" else table.column_count
    if index >= available:
        kind = "行" if read_direction == "单行" else "列"
        raise ValueError(
            f"{kind}编号 {index} 超出范围；当前可用编号为 0 到 {available - 1}。"
        )
    items = table.rows[index] if read_direction == "单行" else tuple(
        row[index] for row in table.rows
    )
    items = items[skip_items:]
    if max_items > 0:
        items = items[:max_items]
    if not items:
        raise ValueError("跳过前 N 项或应用读取数量上限后没有剩余数据。")
    return tuple(items)


def read_excel_table(
    excel_path: str,
    sheet_name: str = "",
    cell_range: str = "",
    skip_rows: int = 0,
    skip_columns: int = 0,
    formula_mode: str = "缓存值",
) -> ExcelTable:
    """Read one worksheet into a cropped, rectangular string table."""

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise RuntimeError(
            "缺少 Excel 读取依赖 openpyxl。请在 ComfyUI 的 Python 环境中安装 requirements.txt。"
        ) from exc

    path = _resolve_excel_path(excel_path)
    if formula_mode not in FORMULA_MODES:
        raise ValueError(f"未知公式读取模式：{formula_mode}")
    skip_rows = _non_negative_int(skip_rows, "跳过行数")
    skip_columns = _non_negative_int(skip_columns, "跳过列数")

    workbook = None
    formula_workbook = None
    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=formula_mode == "缓存值",
            keep_links=False,
        )
        if formula_mode == "缓存值":
            formula_workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
    except (InvalidFileException, OSError, KeyError, ValueError, BadZipFile, EOFError) as exc:
        if workbook is not None:
            workbook.close()
        raise ValueError(f"无法读取 Excel 文件，文件可能已损坏或受密码保护：{path}") from exc

    try:
        worksheets = list(workbook.worksheets)
        sheet_names = tuple(sheet.title for sheet in worksheets)
        if not worksheets:
            raise ValueError("Excel 文件中没有可读取的工作表。")

        requested_sheet = str(sheet_name)
        if requested_sheet != "":
            if requested_sheet not in sheet_names:
                names = "、".join(sheet_names)
                raise ValueError(f"找不到工作表“{requested_sheet}”；可用工作表：{names}")
            worksheet = workbook[requested_sheet]
        else:
            worksheet = worksheets[0]

        bounds = _range_bounds(cell_range)
        if bounds is None:
            min_col, min_row, max_col, max_row = (
                1,
                1,
                max(1, worksheet.max_column),
                max(1, worksheet.max_row),
            )
        else:
            min_col, min_row, max_col, max_row = bounds

        raw_rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
        if formula_workbook is not None:
            formula_worksheet = formula_workbook[worksheet.title]
            formula_rows = [
                list(row)
                for row in formula_worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            ]
            raw_rows = _restore_uncached_formula_cells(raw_rows, formula_rows)
        raw_rows = _crop_outer_empty(raw_rows)
        if not raw_rows:
            raise ValueError("指定工作表或范围内没有可读取的数据。")

        formatted = tuple(tuple(_format_value(value) for value in row) for row in raw_rows)
        table = ExcelTable(formatted, sheet_names, worksheet.title)
        return slice_excel_table(table, skip_rows=skip_rows, skip_columns=skip_columns)
    finally:
        workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()


def table_text_items(table: ExcelTable, orientation: str, cell_separator: str) -> list[str]:
    """Convert a table to one text item per row or per column."""

    if orientation not in ORIENTATIONS:
        raise ValueError(f"未知组织方式：{orientation}")
    separator = decode_separator(cell_separator)
    groups = table.rows if orientation == "按行" else tuple(zip(*table.rows))
    return [separator.join(group) for group in groups]
